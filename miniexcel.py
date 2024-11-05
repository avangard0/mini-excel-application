import openpyxl
import os
import tkinter as tk
from tkinter import messagebox
import re
import statistics

# Имя Excel файла
file_name = 'D:/report.xlsx'

# Функция для создания новой книги Excel с двумя листами: "Сотрудники" и "Товары"
def create_excel_file():
    wb = openpyxl.Workbook()

    # Лист для отчетов сотрудников
    ws1 = wb.active
    ws1.title = 'Сотрудники'
    ws1.append(['ФИО', 'ТабельныйНомер', 'ДатаРождения', 'Возраст', 'Должность', 'Дата_Приема', 'Стаж', 'Оклад'])

    # Лист для товаров
    ws2 = wb.create_sheet(title='Товары')
    ws2.append(['Наименование', 'Количество', 'Стоимость', 'Куплено', 'Окупаемость'])

    wb.save(file_name)

# Функция для проверки наличия заголовков на листе
def check_and_add_headers(ws, headers):
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(headers)

# Функция для проверки корректности ввода даты (формат ДД.ММ.ГГГГ)
def validate_date(date):
    pattern = r'\d{2}\.\d{2}\.\d{4}'
    return re.fullmatch(pattern, date) is not None

# Функция для проверки, что числовые поля введены корректно
def validate_numeric(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

# ===== Функции для работы с сотрудниками 
# Функция для добавления данных сотрудников в Excel
def add_employee_data_to_excel(data):
    try:
        if not os.path.exists(file_name):
            create_excel_file()

        wb = openpyxl.load_workbook(file_name)
        ws = wb['Сотрудники']

        # Проверяем и добавляем заголовки, если их нет
        check_and_add_headers(ws, ['ФИО', 'ТабельныйНомер', 'ДатаРождения', 'Возраст', 'Должность', 'Дата_Приема', 'Стаж', 'Оклад'])

        ws.append(data)
        wb.save(file_name)
        messagebox.showinfo("Успех", "Данные сотрудника успешно сохранены в Excel!")
    except PermissionError:
        messagebox.showerror("Ошибка", "Невозможно записать в файл. Возможно, файл открыт в другой программе.")

# Функция для обработки нажатия кнопки "Добавить данные" для сотрудников
def submit_employee_data():
    fio = entry_fio.get()
    employee_id = entry_employee_id.get()
    birth_date = entry_birth_date.get()
    age = entry_age.get()
    position = entry_position.get()
    hire_date = entry_hire_date.get()
    experience = entry_experience.get()
    salary = entry_salary.get()

    # Проверка на заполненность всех полей
    if (fio == "" or employee_id == "" or birth_date == "" or age == "" or position == "" or 
        hire_date == "" or experience == "" or salary == ""):
        messagebox.showwarning("Ошибка", "Все поля должны быть заполнены!")
    elif not validate_numeric(age):
        messagebox.showwarning("Ошибка", "Возраст должен быть числом!")
    elif not validate_numeric(experience):
        messagebox.showwarning("Ошибка", "Стаж должен быть числом!")
    elif not validate_numeric(salary):
        messagebox.showwarning("Ошибка", "Оклад должен быть числом!")
    elif not validate_date(birth_date):
        messagebox.showwarning("Ошибка", "Неверный формат Даты рождения! Используйте ДД.ММ.ГГГГ.")
    elif not validate_date(hire_date):
        messagebox.showwarning("Ошибка", "Неверный формат Даты приема! Используйте ДД.ММ.ГГГГ.")
    else:
        # Преобразуем возраст, стаж и оклад в числа (целые или с плавающей запятой)
        age = int(age)
        experience = int(experience)
        salary = float(salary)
        employee_id = int(employee_id)
        data = [fio, employee_id, birth_date, age, position, hire_date, experience, salary]
        add_employee_data_to_excel(data)
        clear_employee_entries()

# Функция для отображения данных из Excel
def show_data_from_excel(sheet_name):
    try:
        if not os.path.exists(file_name):
            messagebox.showwarning("Ошибка", "Файл не найден!")
            return

        wb = openpyxl.load_workbook(file_name)
        if sheet_name not in wb.sheetnames:
            messagebox.showwarning("Ошибка", f"Лист {sheet_name} не найден!")
            return

        ws = wb[sheet_name]

        # Создаем новое окно для отображения данных
        data_window = tk.Toplevel(root)
        data_window.title(f"Данные: {sheet_name}")

        # Выводим заголовки
        for col_num, header in enumerate(ws[1], 1):
            label_header = tk.Label(data_window, text=header.value, font=('Arial', 10, 'bold'))
            label_header.grid(row=0, column=col_num - 1, padx=5, pady=5)

        # Выводим строки данных
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            for col_num, value in enumerate(row, 1):
                label_value = tk.Label(data_window, text=value)
                label_value.grid(row=row_num, column=col_num - 1, padx=5, pady=5)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось прочитать данные: {str(e)}")

# Функция для редактирования данных сотрудника по Табельному номеру
def edit_employee_record():
    employee_id = entry_employee_id.get()
    if not employee_id.isdigit():
        messagebox.showwarning("Ошибка", "Введите числовой табельный номер для редактирования!")
        return
    
    employee_id = int(employee_id)
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Сотрудники']
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value == employee_id:
                # Обновляем данные на основе введённых значений
                row[0].value = entry_fio.get()
                row[2].value = entry_birth_date.get()
                row[3].value = int(entry_age.get())
                row[4].value = entry_position.get()
                row[5].value = entry_hire_date.get()
                row[6].value = int(entry_experience.get())
                row[7].value = float(entry_salary.get())
                wb.save(file_name)
                messagebox.showinfo("Успех", "Данные сотрудника успешно обновлены!")
                return
        
        messagebox.showwarning("Ошибка", "Сотрудник с данным табельным номером не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при редактировании: {str(e)}")

# Функция для расчета статистики по зарплате
def salary_statistics():
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Сотрудники']
        
        salaries = [row[7] for row in ws.iter_rows(min_row=2, values_only=True)]
        if not salaries:
            messagebox.showinfo("Результат", "Нет данных о зарплатах сотрудников.")
            return
        
        min_salary = min(salaries)
        max_salary = max(salaries)
        median_salary = statistics.median(salaries)
        
        messagebox.showinfo("Статистика зарплат", f"Минимальная зарплата: {min_salary:.2f}\n"
                                                   f"Максимальная зарплата: {max_salary:.2f}\n"
                                                   f"Медианная зарплата: {median_salary:.2f}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при расчете статистики: {str(e)}")

# Функция для просмотра данных сотрудников
def view_employee_data():
    show_data_from_excel('Сотрудники')

# Функция для просмотра данных товаров
def view_product_data():
    show_data_from_excel('Товары')

# Функция для удаления записи сотрудника по Табельному номеру
def delete_employee_record():
    employee_id = entry_employee_id.get()
    if not employee_id.isdigit():
        messagebox.showwarning("Ошибка", "Введите числовой табельный номер для удаления!")
        return
    
    employee_id = int(employee_id)
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Сотрудники']
        
        # Находим и удаляем строку
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value == employee_id:
                ws.delete_rows(row[0].row, 1)
                wb.save(file_name)
                messagebox.showinfo("Успех", "Запись сотрудника удалена!")
                return
        
        messagebox.showwarning("Ошибка", "Сотрудник с данным табельным номером не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при удалении: {str(e)}")

# Функция для очистки полей ввода сотрудников
def clear_employee_entries():
    entry_fio.delete(0, tk.END)
    entry_employee_id.delete(0, tk.END)
    entry_birth_date.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_position.delete(0, tk.END)
    entry_hire_date.delete(0, tk.END)
    entry_experience.delete(0, tk.END)
    entry_salary.delete(0, tk.END)


# ===== Функции для работы с товарами 
# Функция для добавления данных товаров в Excel
def add_product_data_to_excel(data):
    try:
        if not os.path.exists(file_name):
            create_excel_file()

        wb = openpyxl.load_workbook(file_name)
        ws = wb['Товары']

        # Проверяем и добавляем заголовки, если их нет
        check_and_add_headers(ws, ['Наименование', 'Количество', 'Стоимость', 'Куплено', 'Окупаемость'])

        ws.append(data)
        wb.save(file_name)
        messagebox.showinfo("Успех", "Данные о товаре успешно сохранены в Excel!")
    except PermissionError:
        messagebox.showerror("Ошибка", "Невозможно записать в файл. Возможно, файл открыт в другой программе.")

# Функция для обработки нажатия кнопки "Добавить данные" для товаров
def submit_product_data():
    name = entry_name.get()
    quantity = entry_quantity.get()
    price = entry_price.get()
    purchased = entry_purchased.get()
    roi = entry_roi.get()

    # Проверка на заполненность всех полей
    if name == "" or quantity == "" or price == "" or purchased == "" or roi == "":
        messagebox.showwarning("Ошибка", "Все поля должны быть заполнены!")
    elif not validate_numeric(quantity):
        messagebox.showwarning("Ошибка", "Количество должно быть числом!")
    elif not validate_numeric(price):
        messagebox.showwarning("Ошибка", "Стоимость должна быть числом!")
    elif not validate_numeric(roi):
        messagebox.showwarning("Ошибка", "Окупаемость должна быть числом!")
    else:
        # Преобразуем числовые значения в числа
        quantity = int(quantity)
        price = float(price)
        roi = float(roi)
        purchased = purchased.lower() in ['да', 'yes', 'true']  # Преобразуем "Куплено" в булево значение
        data = [name, quantity, price, purchased, roi]
        add_product_data_to_excel(data)
        clear_product_entries()

# Функция для удаления записи товара по Наименованию
def delete_product_record():
    name = entry_name.get().strip()
    if not name:
        messagebox.showwarning("Ошибка", "Введите наименование товара для удаления!")
        return
    
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Товары']
        
        # Находим и удаляем строку
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == name:
                ws.delete_rows(row[0].row, 1)
                wb.save(file_name)
                messagebox.showinfo("Успех", "Запись товара удалена!")
                return
        
        messagebox.showwarning("Ошибка", "Товар с данным наименованием не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при удалении: {str(e)}")


# Функция для очистки полей ввода товаров
def clear_product_entries():
    entry_name.delete(0, tk.END)
    entry_quantity.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_purchased.delete(0, tk.END)
    entry_roi.delete(0, tk.END)

# Функция для редактирования данных товара по Наименованию
def edit_product_record():
    name = entry_name.get().strip()
    if not name:
        messagebox.showwarning("Ошибка", "Введите наименование товара для редактирования!")
        return
    
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Товары']
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == name:
                # Обновляем данные на основе введённых значений
                row[1].value = int(entry_quantity.get())
                row[2].value = float(entry_price.get())
                row[3].value = entry_purchased.get().lower() in ['да', 'yes', 'true']
                row[4].value = float(entry_roi.get())
                wb.save(file_name)
                messagebox.showinfo("Успех", "Данные товара успешно обновлены!")
                return
        
        messagebox.showwarning("Ошибка", "Товар с данным наименованием не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при редактировании: {str(e)}")

# Функция для удаления записи товара по Наименованию
def delete_product_record():
    name = entry_name.get().strip()
    if not name:
        messagebox.showwarning("Ошибка", "Введите наименование товара для удаления!")
        return
    
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Товары']
        
        # Находим и удаляем строку
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == name:
                ws.delete_rows(row[0].row, 1)
                wb.save(file_name)
                messagebox.showinfo("Успех", "Запись товара удалена!")
                return
        
        messagebox.showwarning("Ошибка", "Товар с данным наименованием не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при удалении: {str(e)}")

# Функция для редактирования данных товара по Наименованию
def edit_product_record():
    name = entry_name.get().strip()
    if not name:
        messagebox.showwarning("Ошибка", "Введите наименование товара для редактирования!")
        return
    
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Товары']
        
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == name:
                # Обновляем данные на основе введённых значений
                row[1].value = int(entry_quantity.get())
                row[2].value = float(entry_price.get())
                row[3].value = entry_purchased.get().lower() in ['да', 'yes', 'true']
                row[4].value = float(entry_roi.get())
                wb.save(file_name)
                messagebox.showinfo("Успех", "Данные товара успешно обновлены!")
                return
        
        messagebox.showwarning("Ошибка", "Товар с данным наименованием не найден!")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при редактировании: {str(e)}")

# Функция для завершения программы
def exit_program():
    root.quit()

# Создаем графический интерфейс с помощью tkinter
root = tk.Tk()
root.title("База данных сотрудников и товаров")

# ===== Интерфейс для сотрудников 
label_employee_section = tk.Label(root, text="Сотрудники", font=('Arial', 12, 'bold'))
label_employee_section.grid(row=0, column=0, columnspan=2, pady=10)

label_fio = tk.Label(root, text="ФИО")
label_fio.grid(row=1, column=0, padx=10, pady=5)

label_employee_id = tk.Label(root, text="Табельный номер")
label_employee_id.grid(row=2, column=0, padx=10, pady=5)

label_birth_date = tk.Label(root, text="Дата рождения (ДД.ММ.ГГГГ)")
label_birth_date.grid(row=3, column=0, padx=10, pady=5)

label_age = tk.Label(root, text="Возраст")
label_age.grid(row=4, column=0, padx=10, pady=5)

label_position = tk.Label(root, text="Должность")
label_position.grid(row=5, column=0, padx=10, pady=5)

label_hire_date = tk.Label(root, text="Дата приема (ДД.ММ.ГГГГ)")
label_hire_date.grid(row=6, column=0, padx=10, pady=5)

label_experience = tk.Label(root, text="Стаж")
label_experience.grid(row=7, column=0, padx=10, pady=5)

label_salary = tk.Label(root, text="Оклад")
label_salary.grid(row=8, column=0, padx=10, pady=5)

# Поля ввода для сотрудников
entry_fio = tk.Entry(root)
entry_fio.grid(row=1, column=1, padx=10, pady=5)

entry_employee_id = tk.Entry(root)
entry_employee_id.grid(row=2, column=1, padx=10, pady=5)

entry_birth_date = tk.Entry(root)
entry_birth_date.grid(row=3, column=1, padx=10, pady=5)

entry_age = tk.Entry(root)
entry_age.grid(row=4, column=1, padx=10, pady=5)

entry_position = tk.Entry(root)
entry_position.grid(row=5, column=1, padx=10, pady=5)

entry_hire_date = tk.Entry(root)
entry_hire_date.grid(row=6, column=1, padx=10, pady=5)

entry_experience = tk.Entry(root)
entry_experience.grid(row=7, column=1, padx=10, pady=5)

entry_salary = tk.Entry(root)
entry_salary.grid(row=8, column=1, padx=10, pady=5)

# Кнопка для добавления данных сотрудников
button_submit_employee = tk.Button(root, text="Добавить данные сотрудника", command=submit_employee_data)
button_submit_employee.grid(row=9, column=0, columnspan=2, pady=10)


# ===== Интерфейс для товаров 
label_product_section = tk.Label(root, text="Товары", font=('Arial', 12, 'bold'))
label_product_section.grid(row=0, column=3, columnspan=2, pady=10)

label_name = tk.Label(root, text="Наименование")
label_name.grid(row=1, column=3, padx=10, pady=5)

label_quantity = tk.Label(root, text="Количество")
label_quantity.grid(row=2, column=3, padx=10, pady=5)

label_price = tk.Label(root, text="Стоимость")
label_price.grid(row=3, column=3, padx=10, pady=5)

label_purchased = tk.Label(root, text="Куплено (Да/Нет)")
label_purchased.grid(row=4, column=3, padx=10, pady=5)

label_roi = tk.Label(root, text="Окупаемость")
label_roi.grid(row=5, column=3, padx=10, pady=5)

# Поля ввода для товаров
entry_name = tk.Entry(root)
entry_name.grid(row=1, column=4, padx=10, pady=5)

entry_quantity = tk.Entry(root)
entry_quantity.grid(row=2, column=4, padx=10, pady=5)

entry_price = tk.Entry(root)
entry_price.grid(row=3, column=4, padx=10, pady=5)

entry_purchased = tk.Entry(root)
entry_purchased.grid(row=4, column=4, padx=10, pady=5)

entry_roi = tk.Entry(root)
entry_roi.grid(row=5, column=4, padx=10, pady=5)

# Кнопка для добавления данных товаров
button_submit_product = tk.Button(root, text="Добавить данные товара", command=submit_product_data)
button_submit_product.grid(row=6, column=3, columnspan=2, pady=10)

# Добавляем кнопки для просмотра 
button_view_employee = tk.Button(root, text="Посмотреть данные сотрудников", command=view_employee_data)
button_view_employee.grid(row=11, column=0, columnspan=2, pady=10)

button_view_product = tk.Button(root, text="Посмотреть данные товаров", command=view_product_data)
button_view_product.grid(row=8, column=3, columnspan=2, pady=10)

button_delete_employee = tk.Button(root, text="Удалить сотрудника по табельному номеру", command=delete_employee_record)
button_delete_employee.grid(row=12, column=0, columnspan=2, pady=10)

button_delete_product = tk.Button(root, text="Удалить товар по наименованию", command=delete_product_record)
button_delete_product.grid(row=9, column=3, columnspan=2, pady=10)

button_edit_employee = tk.Button(root, text="Редактировать данные сотрудника", command=edit_employee_record)
button_edit_employee.grid(row=10, column=0, columnspan=2, pady=10)

button_edit_product = tk.Button(root, text="Редактировать данные товара", command=edit_product_record)
button_edit_product.grid(row=7, column=3,columnspan=2, pady=10)

button_salary_statistics = tk.Button(root, text="Статистика по зарплате сотрудников", command=salary_statistics)
button_salary_statistics.grid(row=13, column=0,columnspan=2, pady=10)

# Кнопка для выхода
button_exit = tk.Button(root, text="Выход", command=exit_program)
button_exit.grid(row=11, column=4, columnspan = 2, pady=10)

# Запуск приложения
root.mainloop()